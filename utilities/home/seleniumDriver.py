############################
### BY: HARPREET DHILLON ###
############################

############################
#####  #####
############################
import datetime
import os

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from traceback import print_stack
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
import utilities.home.custom_logger as cl
import logging
from time import sleep, strftime
import pdb

from utilities.home.screenshots import ss


class SeleniumDriver:
    log = cl.customLogger(logging.DEBUG)

    def __init__(self, driver):
        self.driver = driver
    # Getting the first column from the sheet to give the name while taking screenshot

    # Update the result column
    def resultColumn(self, count, result):
        dirPath = os.getcwd()
        filepath = f"{dirPath}\\Data.xlsx"
        wb = load_workbook(filepath)
        sheet = wb["MDCO-E2E"]
        sheet[f'AH{count}'] = result
        # print(f"{count}: {result}")
        wb.save(filepath)

    # Added the PDCO-E2E Sheet data into the Dictionary to access that data later
    def sheetData(self):
        # set file path
        # filepath = "F://Digital//PycharmProjects//PDCO//Data.xlsx"
        filepath = "F://Working//Pycharm//PDCO-DigitalE2E//Data.xlsx"
        # load .xlsx
        wb = load_workbook(filepath)
        sheet = wb["MDCO-E2E"]

        maxCol = sheet.max_column

        userData = {}
        columns = []
        for col in sheet.iter_cols(min_row=1,
                                   min_col=1,
                                   max_col=maxCol,
                                   values_only=True):
            columnName = col[0]
            columns.append(columnName)
        for row in sheet.iter_rows(min_row=2,
                                   min_col=1,
                                   max_col=maxCol,
                                   values_only=True):
            product_id = row[0]
            product = {
                columns[1]: row[1],
                columns[2]: row[2],
                columns[3]: row[3],
                columns[4]: row[4],
                columns[5]: row[5],
                columns[6]: row[6],
                columns[7]: row[7],
                columns[8]: row[8],
                columns[9]: row[9],
                columns[10]: row[10],
                columns[11]: row[11],
                columns[12]: row[12],
                columns[13]: row[13],
                columns[14]: row[14],
                columns[15]: row[15],
                columns[16]: row[16],
                columns[17]: row[17],
                columns[18]: row[18],
                columns[19]: row[19],
                columns[20]: row[20],
                columns[21]: row[21],
                columns[22]: row[22],
                columns[23]: row[23],
                columns[24]: row[24],
                columns[25]: row[25],
                columns[26]: row[26],
                columns[27]: row[27],
                columns[28]: row[28],
                columns[29]: row[29],
                columns[30]: row[30],
                columns[31]: row[31],
                columns[32]: row[32],
                columns[33]: row[33]
            }
            userData[product_id] = product
        return userData

    #######################
    ### Compare 2 lists ###
    #######################
    def compareList(self, expectedList, actualList, tfsNumber, count):
        for i in range(len(actualList)):
            # if actualList[i] == expectedList[i]:
            #     self.takeScreenshot(self.driver, f"PASSED_{tfsNumber}")
            #     self.resultColumn(count, "PASSED")
            if actualList[i] != expectedList[i]:
                self.takeScreenshot(self.driver, f"FAILED_{tfsNumber}")
                self.resultColumn(count, "FAILED")
            # else:
            #     self.takeScreenshot(self.driver, f"FAILED_{tfsNumber}")
            #     self.resultColumn(count, "FAILED")
            assert actualList[i] == expectedList[i], f"Check value in the list \n" \
                                                     f"Expected: {expectedList[i]}\nActual: {actualList[i]}"

    def getByType(self, locatorType):
        locatorType = locatorType.lower()
        if locatorType == "id":
            return By.ID
        elif locatorType == "name":
            return By.NAME
        elif locatorType == "xpath":
            return By.XPATH
        elif locatorType == "css":
            return By.CSS_SELECTOR
        elif locatorType == "class":
            return By.CLASS_NAME
        elif locatorType == "link":
            return By.LINK_TEXT
        elif locatorType == 'tag':
            return By.TAG_NAME
        else:
            self.log.info("Locator type " + locatorType +
                          " not correct/supported")
        return False

    def getElement(self, locator, locatorType="id"):
        element = None
        try:
            locatorType = locatorType.lower()
            byType = self.getByType(locatorType)
            self.waitForElement(locator, locatorType)
            element = self.driver.find_element(byType, locator)
            # print("getElement: Element Found ")
            # print(element)
        except:
            self.log.info("Element not found with locator: " + locator +
                       " and  locatorType: " + locatorType)
            # print("getElement: Element NOT Found ")
            # print(element)
        return element

    def elementClick(self, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)
            element.click()
            self.log.info("Clicked on element with locator: " + locator +
                          " locatorType: " + locatorType)
        except:
            self.log.info("Cannot click on the element with locator: " + locator +
                          " locatorType: " + locatorType)
            print_stack()

    def sendKeys(self, data, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)
            element.send_keys(data)
            self.log.info("Sent data on element with locator: " + locator +
                          " locatorType: " + locatorType)
        except:
            self.log.info("Cannot send data on the element with locator: " + locator +
                          " locatorType: " + locatorType)
            print_stack()

    def isElementPresent(self, locator, locatorType="id"):
        try:
            element = self.getElement(locator, locatorType)
            if element is not None:
                self.log.info("Element Found")
                return False
            else:
                self.log.info("Element not found")
                return False
        except:
            print("Element not found")
            return False

    def elementsPresenceCheck(self, locator, byType):
        try:
            elementList = self.driver.find_elements(byType, locator)
            if len(elementList) > 0:
                self.log.info("Element Found")
                return True
            else:
                self.log.info("Element not found")
                return False
        except:
            self.log.info("Element not found")
            return False

    def waitForElement(self, locator, locatorType="id",
                       timeout=5, pollFrequency=0.5):
        element = None
        try:
            byType = self.getByType(locatorType)
            self.log.info("Waiting for maximum :: " + str(timeout) +
                          " :: seconds for element to be located")
            wait = WebDriverWait(self.driver, timeout=timeout, poll_frequency=pollFrequency,
                                 ignored_exceptions=[NoSuchElementException,
                                                     ElementNotVisibleException,
                                                     ElementNotSelectableException])
            element = wait.until(EC.presence_of_element_located((byType, locator)))
        except:
            print(f"Element at {locator} not appeared on the web page")
        return element

    def takeScreenshot(self, driver, product_id):
        """
        Takes screenshot of the current open web page
        """
        fileName = str(product_id) + strftime('_%m%d%Y_%H%M%S') + ".png"
        dirPath = os.getcwd()
        screenshotDirectory = f"{dirPath}\\screenshots\\"
        # screenshotDirectory = "F://Working//Pycharm//MDCO-DgitalE2E//screenshots//"
        destinationFile = screenshotDirectory + fileName

        try:
            sleep(3)
            driver.save_screenshot(destinationFile)
            print("Screenshot saved to directory --> :: " + destinationFile)
        except NotADirectoryError:
            print("Not a directory issue")
